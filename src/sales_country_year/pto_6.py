# 6. Access and display the value of cell B2 (the data source for the workbook)

from openpyxl import load_workbook

# Especifica la ruta del archivo .xlsx
archivo_excel = 'D:/dev/tp_iglesias_2/files/EcarSalesByCountryAndYear2.xlsx'

# Cargar el libro de trabajo usando load_workbook() y asignarlo a la variable 'workbook'
workbook = load_workbook(filename=archivo_excel)

# Acceder a la pestaña 'Documentación'
documentacion_sheet = workbook['Documentación']

# Acceder al valor de la celda B2 (fuente de datos)
fuente_datos = documentacion_sheet['B2'].value

# Mostrar el valor de la celda B2
print(f"Fuente de datos: {fuente_datos}")
