# 5. Access the 'Documentation' worksheet tab using workbook[ ]

from openpyxl import load_workbook

# Especifica la ruta del archivo .xlsx
archivo_excel = 'D:/dev/tp_iglesias_2/files/EcarSalesByCountryAndYear2.xlsx'

# Cargar el libro de trabajo usando load_workbook() y asignarlo a la variable 'workbook'
workbook = load_workbook(filename=archivo_excel)

# Acceder a la pestaña 'Documentación'
documentacion_sheet = workbook['Documentación']

# Mostrar el título de la hoja para verificar
print(f"Título de la hoja: {documentacion_sheet.title}")
