# 4. Use the openpyxl load_workbook() function to load the workbook, assigning the result to a variable called workbook

from openpyxl import load_workbook

# Especifica la ruta del archivo .xlsx
archivo_excel = 'D:/dev/tp_iglesias_2/files/EcarSalesByCountryAndYear2.xlsx'

# Cargar el libro de trabajo usando load_workbook() y asignarlo a la variable 'workbook'
workbook = load_workbook(filename=archivo_excel)

# Mostrar las hojas de trabajo disponibles en el libro
print(workbook.sheetnames)
